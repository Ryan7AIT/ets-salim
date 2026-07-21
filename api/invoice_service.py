"""Invoice CRUD, settings, and export helpers."""

from __future__ import annotations

import base64
import io
import re
import sqlite3
from datetime import date, datetime, timedelta, timezone
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

try:
    from .notification_service import read_json_setting, write_json_setting
except ImportError:
    from notification_service import read_json_setting, write_json_setting


INVOICE_SETTINGS_KEY = "invoiceSettings"
INVOICES_ENABLED_ENV = "INVOICES_ENABLED"

DEFAULT_INVOICE_SETTINGS: dict[str, Any] = {
    "companyName": "Ets Bellal Salim",
    "contactName": "",
    "address": "",
    "email": "",
    "phone": "",
    "nif": "",
    "registrationNumber": "",
    "rip": "",
    "logoMode": "text",
    "logoText": "Ets Bellal Salim",
    "logoImage": "",
    "cachetImage": "",
    "invoiceLanguage": "fr",
    "defaultTaxRate": 20,
    "currency": "DZD",
    "paymentTermsDays": 7,
    "footerNotes": "",
}

INVOICE_LABELS: dict[str, dict[str, str]] = {
    "en": {
        "title": "Invoice",
        "proformaTitle": "Proforma",
        "invoiceNumber": "Invoice Number:",
        "proformaNumber": "Proforma Number:",
        "issueDate": "Date of Issue:",
        "dueDate": "Date Due:",
        "seller": "From:",
        "billTo": "Bill To:",
        "client": "Client:",
        "description": "Description",
        "quantity": "Qty",
        "unitPrice": "Unit price",
        "amount": "Amount",
        "notes": "Notes:",
        "subtotal": "Subtotal",
        "adjustments": "Adjustments",
        "discount": "Discount",
        "adjustedSubtotal": "Adjusted Subtotal",
        "tax": "Tax",
        "total": "Total",
        "nif": "NIF:",
        "rc": "R.C:",
        "nis": "NIS:",
        "registrationNumber": "Registration No.:",
        "rip": "RIP:",
    },
    "fr": {
        "title": "Facture",
        "proformaTitle": "Proforma",
        "invoiceNumber": "Numéro de facture :",
        "proformaNumber": "Numéro proforma :",
        "issueDate": "Date d'émission :",
        "dueDate": "Date d'échéance :",
        "seller": "Émetteur :",
        "billTo": "Facturé à :",
        "client": "Client :",
        "description": "Description",
        "quantity": "Qté",
        "unitPrice": "Prix unitaire",
        "amount": "Montant",
        "notes": "Notes :",
        "subtotal": "Sous-total",
        "adjustments": "Ajustements",
        "discount": "Remise",
        "adjustedSubtotal": "Sous-total ajusté",
        "tax": "TVA",
        "total": "Total",
        "nif": "NIF :",
        "rc": "R.C :",
        "nis": "NIS :",
        "registrationNumber": "N° d'immatriculation :",
        "rip": "RIP :",
    },
}


def env_flag_enabled(value: str | None) -> bool:
    if not value:
        return False
    return value.strip().lower() in {"1", "true", "yes", "on"}


def invoices_enabled(read_config_value) -> bool:
    return env_flag_enabled(read_config_value(INVOICES_ENABLED_ENV))


def require_invoices_enabled(read_config_value) -> None:
    if not invoices_enabled(read_config_value):
        raise HTTPException(status_code=404, detail="Invoice module is not enabled")


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def default_invoice_settings() -> dict[str, Any]:
    return dict(DEFAULT_INVOICE_SETTINGS)


def invoice_labels(settings: dict[str, Any]) -> dict[str, str]:
    language = (settings.get("invoiceLanguage") or "fr").lower()
    return INVOICE_LABELS.get(language, INVOICE_LABELS["fr"])


def is_proforma(invoice: dict[str, Any]) -> bool:
    return (invoice.get("documentType") or "facture").lower() == "proforma"


def document_labels(invoice: dict[str, Any], settings: dict[str, Any]) -> dict[str, str]:
    labels = invoice_labels(settings)
    if is_proforma(invoice):
        return {
            **labels,
            "title": labels.get("proformaTitle", "Proforma"),
            "invoiceNumber": labels.get("proformaNumber", labels["invoiceNumber"]),
        }
    return labels


def load_invoice_settings(conn: sqlite3.Connection) -> dict[str, Any]:
    stored = read_json_setting(conn, INVOICE_SETTINGS_KEY, default_invoice_settings())
    merged = default_invoice_settings()
    merged.update(stored or {})
    if merged.get("invoiceLanguage") not in INVOICE_LABELS:
        merged["invoiceLanguage"] = "fr"
    return merged


def save_invoice_settings(conn: sqlite3.Connection, settings: dict[str, Any]) -> dict[str, Any]:
    merged = default_invoice_settings()
    merged.update(settings)
    write_json_setting(conn, INVOICE_SETTINGS_KEY, merged)
    return merged


def line_amount(quantity: float, unit_price: float) -> float:
    return round(float(quantity or 0) * float(unit_price or 0), 2)


def compute_totals(
    items: list[dict[str, Any]],
    adjustment: float,
    tax_rate: float,
    discount_amount: float = 0,
) -> dict[str, float]:
    subtotal = round(sum(line_amount(item.get("quantity", 0), item.get("unitPrice", 0)) for item in items), 2)
    adjustment = round(float(adjustment or 0), 2)
    discount = round(max(0, float(discount_amount or 0)), 2)
    adjusted_subtotal = round(subtotal + adjustment, 2)
    taxable_subtotal = round(max(0, adjusted_subtotal - discount), 2)
    tax = round(taxable_subtotal * float(tax_rate or 0) / 100, 2)
    total = round(taxable_subtotal + tax, 2)
    return {
        "subtotal": subtotal,
        "adjustment": adjustment,
        "discount": discount,
        "adjustedSubtotal": adjusted_subtotal,
        "taxableSubtotal": taxable_subtotal,
        "tax": tax,
        "total": total,
    }


def format_money(amount: float, currency: str = "DZD") -> str:
    symbol = {"DZD": "DA", "EUR": "€", "USD": "$"}.get(currency.upper(), currency)
    if symbol in {"€", "$"}:
        return f"{symbol}{amount:,.2f}"
    return f"{amount:,.2f} {symbol}"


def format_display_date(value: str) -> str:
    if not value:
        return ""
    try:
        return date.fromisoformat(value).strftime("%d/%m/%Y")
    except ValueError:
        return value


def format_client_code(client: dict[str, Any]) -> str:
    client_id = client.get("id")
    try:
        return f"C{int(client_id):03d}"
    except (TypeError, ValueError):
        return "C---"


def next_invoice_number(conn: sqlite3.Connection) -> str:
    row = conn.execute("SELECT number FROM invoices ORDER BY id DESC LIMIT 1").fetchone()
    if not row:
        return "1"
    match = re.search(r"(\d+)$", row["number"] or "")
    if not match:
        return str(conn.execute("SELECT COUNT(*) FROM invoices").fetchone()[0] + 1)
    return str(int(match.group(1)) + 1)


def fetch_invoice_items(conn: sqlite3.Connection, invoice_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT id, description, quantity, unit_price, sort_order
        FROM invoice_items
        WHERE invoice_id = ?
        ORDER BY sort_order, id
        """,
        (invoice_id,),
    ).fetchall()
    items = []
    for row in rows:
        quantity = float(row["quantity"] or 0)
        unit_price = float(row["unit_price"] or 0)
        items.append(
            {
                "id": row["id"],
                "description": row["description"],
                "quantity": quantity,
                "unitPrice": unit_price,
                "amount": line_amount(quantity, unit_price),
            }
        )
    return items


def serialize_invoice_row(conn: sqlite3.Connection, row: sqlite3.Row, include_items: bool = True) -> dict[str, Any]:
    items = fetch_invoice_items(conn, row["id"]) if include_items else []
    totals = compute_totals(items, row["adjustment"], row["tax_rate"], row["discount_amount"])
    client = conn.execute("SELECT * FROM clients WHERE id = ?", (row["client_id"],)).fetchone()
    return {
        "id": row["id"],
        "number": row["number"],
        "clientId": row["client_id"],
        "client": {
            "id": client["id"],
            "company": client["company"],
            "contact": client["contact"],
            "phone": client["phone"] or "",
            "email": client["email"] or "",
            "address": client["address"] or "",
            "nif": client["nif"] or "",
            "rc": client["rc"] or "",
            "nis": client["nis"] or "",
        }
        if client
        else None,
        "issueDate": row["issue_date"],
        "dueDate": row["due_date"] or "",
        "documentType": row["document_type"] if "document_type" in row.keys() else "facture",
        "includeCachet": bool(row["include_cachet"]) if "include_cachet" in row.keys() else False,
        "status": row["status"],
        "currency": row["currency"],
        "notes": row["notes"] or "",
        "adjustment": float(row["adjustment"] or 0),
        "discountAmount": float(row["discount_amount"] or 0),
        "taxRate": float(row["tax_rate"] or 0),
        "items": items,
        "totals": totals,
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
    }


def get_invoice(conn: sqlite3.Connection, invoice_id: int) -> dict[str, Any]:
    row = conn.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return serialize_invoice_row(conn, row)


def list_invoices(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = conn.execute("SELECT * FROM invoices ORDER BY issue_date DESC, id DESC").fetchall()
    return [serialize_invoice_row(conn, row) for row in rows]


def validate_client_exists(conn: sqlite3.Connection, client_id: int) -> sqlite3.Row:
    client = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    if not client:
        raise HTTPException(status_code=400, detail="Client not found")
    return client


def validate_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not items:
        raise HTTPException(status_code=400, detail="At least one line item is required")
    normalized = []
    for index, item in enumerate(items):
        description = (item.get("description") or "").strip()
        if not description:
            raise HTTPException(status_code=400, detail=f"Line item {index + 1} requires a description")
        normalized.append(
            {
                "description": description,
                "quantity": max(0, float(item.get("quantity") or 0)),
                "unitPrice": float(item.get("unitPrice") or 0),
                "sortOrder": int(item.get("sortOrder", index)),
            }
        )
    return normalized


def replace_invoice_items(conn: sqlite3.Connection, invoice_id: int, items: list[dict[str, Any]]) -> None:
    conn.execute("DELETE FROM invoice_items WHERE invoice_id = ?", (invoice_id,))
    for index, item in enumerate(items):
        conn.execute(
            """
            INSERT INTO invoice_items (invoice_id, description, quantity, unit_price, sort_order)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                invoice_id,
                item["description"],
                item["quantity"],
                item["unitPrice"],
                item.get("sortOrder", index),
            ),
        )


def create_invoice(conn: sqlite3.Connection, payload: dict[str, Any]) -> dict[str, Any]:
    validate_client_exists(conn, int(payload["clientId"]))
    items = validate_items(payload.get("items", []))
    now = utc_now_iso()
    issue_date = payload.get("issueDate") or date.today().isoformat()
    settings = load_invoice_settings(conn)
    document_type = (payload.get("documentType") or "facture").lower()
    if document_type not in {"facture", "proforma"}:
        document_type = "facture"
    due_date = payload.get("dueDate")
    if document_type == "proforma":
        due_date = ""
    elif not due_date:
        due_date = (date.fromisoformat(issue_date) + timedelta(days=int(settings.get("paymentTermsDays") or 7))).isoformat()

    number = (payload.get("number") or "").strip() or next_invoice_number(conn)
    cursor = conn.execute(
        """
        INSERT INTO invoices
        (number, client_id, issue_date, due_date, document_type, include_cachet, status, currency, notes, adjustment, discount_amount, tax_rate, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            number,
            int(payload["clientId"]),
            issue_date,
            due_date,
            document_type,
            1 if payload.get("includeCachet") else 0,
            payload.get("status") or "draft",
            payload.get("currency") or settings.get("currency") or "DZD",
            payload.get("notes") or "",
            float(payload.get("adjustment") or 0),
            float(payload.get("discountAmount") or 0),
            float(payload.get("taxRate") if payload.get("taxRate") is not None else settings.get("defaultTaxRate") or 0),
            now,
            now,
        ),
    )
    invoice_id = int(cursor.lastrowid)
    replace_invoice_items(conn, invoice_id, items)
    return get_invoice(conn, invoice_id)


def update_invoice(conn: sqlite3.Connection, invoice_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    row = conn.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Invoice not found")

    validate_client_exists(conn, int(payload["clientId"]))
    items = validate_items(payload.get("items", []))
    now = utc_now_iso()
    document_type = (payload.get("documentType") or row["document_type"] if "document_type" in row.keys() else "facture").lower()
    if document_type not in {"facture", "proforma"}:
        document_type = "facture"
    due_date = payload.get("dueDate")
    if document_type == "proforma":
        due_date = ""
    elif due_date is None:
        due_date = row["due_date"] or ""
    conn.execute(
        """
        UPDATE invoices
        SET number = ?, client_id = ?, issue_date = ?, due_date = ?, document_type = ?, include_cachet = ?, status = ?, currency = ?,
            notes = ?, adjustment = ?, discount_amount = ?, tax_rate = ?, updated_at = ?
        WHERE id = ?
        """,
        (
            (payload.get("number") or row["number"]).strip(),
            int(payload["clientId"]),
            payload.get("issueDate") or row["issue_date"],
            due_date,
            document_type,
            1 if payload.get("includeCachet") else 0,
            payload.get("status") or row["status"],
            payload.get("currency") or row["currency"],
            payload.get("notes") or "",
            float(payload.get("adjustment") if payload.get("adjustment") is not None else row["adjustment"] or 0),
            float(payload.get("discountAmount") if payload.get("discountAmount") is not None else row["discount_amount"] or 0),
            float(payload.get("taxRate") if payload.get("taxRate") is not None else row["tax_rate"] or 0),
            now,
            invoice_id,
        ),
    )
    replace_invoice_items(conn, invoice_id, items)
    return get_invoice(conn, invoice_id)


def delete_invoice(conn: sqlite3.Connection, invoice_id: int) -> None:
    row = conn.execute("SELECT id FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Invoice not found")
    conn.execute("DELETE FROM invoice_items WHERE invoice_id = ?", (invoice_id,))
    conn.execute("DELETE FROM invoices WHERE id = ?", (invoice_id,))


def build_cachet_flowable(settings: dict[str, Any]) -> Any | None:
    if not settings.get("cachetImage"):
        return None
    image_bytes = decode_logo_image(settings["cachetImage"])
    if not image_bytes:
        return None
    return Image(io.BytesIO(image_bytes), width=38 * mm, height=38 * mm, kind="proportional")


def build_logo_flowable(settings: dict[str, Any], logo_text_style: ParagraphStyle) -> Any:
    if settings.get("logoMode") == "image" and settings.get("logoImage"):
        image_bytes = decode_logo_image(settings["logoImage"])
        if image_bytes:
            return Image(io.BytesIO(image_bytes), width=56 * mm, height=30 * mm, kind="proportional")
    return Paragraph(settings.get("logoText") or settings.get("companyName") or "", logo_text_style)


def seller_metadata_lines(settings: dict[str, Any], labels: dict[str, str]) -> list[str]:
    lines = [
        settings.get("companyName") or "",
        settings.get("address") or "",
        settings.get("email") or "",
        settings.get("phone") or "",
    ]
    if settings.get("nif"):
        lines.append(f"{labels['nif']} {settings['nif']}")
    if settings.get("registrationNumber"):
        lines.append(f"{labels['registrationNumber']} {settings['registrationNumber']}")
    if settings.get("rip"):
        lines.append(f"{labels['rip']} {settings['rip']}")
    return [line for line in lines if line]


def proforma_client_paragraphs(client: dict[str, Any], labels: dict[str, str], value_style: ParagraphStyle) -> list[Any]:
    lines: list[Any] = [
        Paragraph(f"<b>{labels['client']}</b>", value_style),
        Paragraph(client.get("company") or "", value_style),
    ]
    if client.get("nif"):
        lines.append(Paragraph(f"{labels['nif']} {client['nif']}", value_style))
    if client.get("rc"):
        lines.append(Paragraph(f"{labels['rc']} {client['rc']}", value_style))
    if client.get("nis"):
        lines.append(Paragraph(f"{labels['nis']} {client['nis']}", value_style))
    return lines


def decode_logo_image(logo_image: str) -> bytes | None:
    if not logo_image:
        return None
    data = logo_image.strip()
    if data.startswith("data:"):
        _, _, data = data.partition(",")
    try:
        return base64.b64decode(data)
    except (ValueError, TypeError):
        return None


def build_pdf(invoice: dict[str, Any], settings: dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    label_style = ParagraphStyle("Label", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748B"))
    value_style = ParagraphStyle("Value", parent=styles["Normal"], fontSize=10, leading=14)
    meta_style = ParagraphStyle(
        "MetaRight",
        parent=value_style,
        fontSize=9.5,
        leading=14,
        alignment=2,
    )
    meta_label_style = ParagraphStyle("MetaLabel", parent=meta_style, fontName="Helvetica-Bold", textColor=colors.HexColor("#334155"))
    meta_compact_style = ParagraphStyle(
        "MetaCompact",
        parent=meta_style,
        fontSize=9,
        textColor=colors.HexColor("#475569"),
    )
    logo_text_style = ParagraphStyle(
        "LogoText",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=26,
        leading=30,
        textColor=colors.HexColor("#0F172A"),
        spaceBefore=4,
    )

    labels = document_labels(invoice, settings)
    story: list[Any] = []
    issue_date = format_display_date(invoice["issueDate"])
    due_date = format_display_date(invoice["dueDate"])
    proforma = is_proforma(invoice)
    header_left = Table(
        [[build_logo_flowable(settings, logo_text_style)]],
        colWidths=[88 * mm],
    )
    header_left.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    meta_rows = [
        [Paragraph(labels["invoiceNumber"], meta_label_style), Paragraph(str(invoice["number"]), meta_style)],
        [Paragraph(labels["issueDate"], meta_label_style), Paragraph(issue_date, meta_compact_style)],
    ]
    if not proforma:
        meta_rows.append(
            [Paragraph(labels["dueDate"], meta_label_style), Paragraph(due_date, meta_compact_style)]
        )
    header_right = Table(meta_rows, colWidths=[35 * mm, 30 * mm])
    header_right.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#E2E8F0")),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 1), (-1, 1), 6),
            ]
        )
    )
    header_table = Table([[header_left, header_right]], colWidths=[100 * mm, 70 * mm])
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                ("LINEBELOW", (0, 0), (-1, 0), 0.6, colors.HexColor("#E2E8F0")),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 14),
                ("TOPPADDING", (0, 0), (-1, 0), 0),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 14))

    client = invoice.get("client") or {}
    ship_lines = [Paragraph(f"<b>{labels['seller']}</b>", value_style)] + [
        Paragraph(line, value_style) for line in seller_metadata_lines(settings, labels)
    ]
    if proforma:
        client_lines = proforma_client_paragraphs(client, labels, value_style)
        address_table = Table([[ship_lines, client_lines]], colWidths=[85 * mm, 85 * mm])
    else:
        bill_lines = [
            Paragraph(f"<b>{labels['billTo']}</b>", value_style),
            Paragraph(f"{format_client_code(client)} - {client.get('company', '')}", value_style),
            Paragraph(client.get("contact") or "", value_style),
            Paragraph(client.get("address") or "", value_style),
            Paragraph(client.get("email") or "", value_style),
        ]
        if client.get("nif"):
            bill_lines.append(Paragraph(f"{labels['nif']} {client['nif']}", value_style))
        if client.get("rc"):
            bill_lines.append(Paragraph(f"{labels['rc']} {client['rc']}", value_style))
        if client.get("nis"):
            bill_lines.append(Paragraph(f"{labels['nis']} {client['nis']}", value_style))
        address_table = Table([[ship_lines, bill_lines]], colWidths=[85 * mm, 85 * mm])
    address_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LINEABOVE", (0, 0), (-1, 0), 0.4, colors.HexColor("#E2E8F0")),
                ("LINEBELOW", (0, 0), (-1, 0), 0.4, colors.HexColor("#E2E8F0")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(address_table)
    story.append(Spacer(1, 12))

    currency = invoice.get("currency") or settings.get("currency") or "DZD"
    item_rows = [[labels["description"], labels["quantity"], labels["unitPrice"], labels["amount"]]]
    for item in invoice.get("items", []):
        item_rows.append(
            [
                item["description"],
                f"{item['quantity']:.0f}" if float(item["quantity"]).is_integer() else f"{item['quantity']:.2f}",
                format_money(item["unitPrice"], currency),
                format_money(item["amount"], currency),
            ]
        )
    while len(item_rows) < 6:
        item_rows.append(["", "", "", ""])

    items_table = Table(item_rows, colWidths=[82 * mm, 18 * mm, 32 * mm, 32 * mm], repeatRows=1)
    items_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(items_table)
    story.append(Spacer(1, 14))

    totals = invoice.get("totals") or {}
    notes_text = invoice.get("notes") or settings.get("footerNotes") or ""
    notes_block = Table(
        [
            [Paragraph(f"<b>{labels['notes']}</b>", value_style)],
            [Paragraph(notes_text.replace("\n", "<br/>"), label_style)],
        ],
        colWidths=[95 * mm],
    )
    notes_block.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    totals_rows = [
        [labels["subtotal"], format_money(totals.get("subtotal", 0), currency)],
        [labels["adjustments"], format_money(totals.get("adjustment", 0), currency)],
        [labels["discount"], f"-{format_money(totals.get('discount', 0), currency)}"],
        [labels["adjustedSubtotal"], format_money(totals.get("taxableSubtotal", 0), currency)],
        [f"{labels['tax']} ({invoice.get('taxRate', 0):g}%)", format_money(totals.get("tax", 0), currency)],
        [labels["total"], format_money(totals.get("total", 0), currency)],
    ]
    totals_table = Table(totals_rows, colWidths=[38 * mm, 32 * mm])
    totals_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (0, -2), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, -1), (-1, -1), 11),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    include_cachet = bool(invoice.get("includeCachet")) and settings.get("cachetImage")
    cachet_flowable = build_cachet_flowable(settings) if include_cachet else None
    if cachet_flowable:
        footer_right = Table([[totals_table], [cachet_flowable]], colWidths=[70 * mm])
        footer_right.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 1), (-1, 1), 22),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
    else:
        footer_right = totals_table

    footer_table = Table([[notes_block, footer_right]], colWidths=[95 * mm, 75 * mm])
    footer_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("ALIGN", (1, 0), (1, 0), "RIGHT")]))
    story.append(footer_table)

    doc.build(story)
    return buffer.getvalue()


def build_excel(invoice: dict[str, Any], settings: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F1F5F9")
    right = Alignment(horizontal="right")
    currency = invoice.get("currency") or settings.get("currency") or "DZD"
    client = invoice.get("client") or {}
    totals = invoice.get("totals") or {}
    labels = document_labels(invoice, settings)
    proforma = is_proforma(invoice)

    issue_date = format_display_date(invoice["issueDate"])
    due_date = format_display_date(invoice["dueDate"])

    sheet["A1"] = settings.get("logoText") or settings.get("companyName") or ""
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["D1"] = f"{labels['invoiceNumber']} {invoice['number']}"
    sheet["D1"].font = Font(size=11, bold=True)
    sheet["D1"].alignment = Alignment(horizontal="right")
    sheet.merge_cells("D1:E1")
    date_line = f"{labels['issueDate']} {issue_date}"
    if not proforma:
        date_line += f"  |  {labels['dueDate']} {due_date}"
    sheet["D2"] = date_line
    sheet["D2"].font = Font(size=10)
    sheet["D2"].alignment = Alignment(horizontal="right")
    sheet.merge_cells("D2:E2")

    sheet.append([])

    if proforma:
        sheet.append([labels["seller"], "", "", labels["client"]])
        sheet.append([settings.get("companyName", ""), "", "", client.get("company", "")])
        client_identifiers = []
        if client.get("nif"):
            client_identifiers.append(f"{labels['nif']} {client['nif']}")
        if client.get("rc"):
            client_identifiers.append(f"{labels['rc']} {client['rc']}")
        if client.get("nis"):
            client_identifiers.append(f"{labels['nis']} {client['nis']}")
        sheet.append([settings.get("address", ""), "", "", client_identifiers[0] if client_identifiers else ""])
        sheet.append([settings.get("email", ""), "", "", client_identifiers[1] if len(client_identifiers) > 1 else ""])
        sheet.append([settings.get("phone", ""), "", "", client_identifiers[2] if len(client_identifiers) > 2 else ""])
    else:
        sheet.append([labels["seller"], "", "", labels["billTo"]])
        sheet.append([settings.get("companyName", ""), "", "", f"{format_client_code(client)} - {client.get('company', '')}"])
        sheet.append([settings.get("address", ""), "", "", client.get("address", "")])
        sheet.append([settings.get("email", ""), "", "", client.get("email", "")])
        sheet.append([settings.get("phone", ""), "", "", client.get("phone", "")])
        if client.get("nif"):
            sheet.append(["", "", "", f"{labels['nif']} {client['nif']}"])
        if client.get("rc"):
            sheet.append(["", "", "", f"{labels['rc']} {client['rc']}"])
        if client.get("nis"):
            sheet.append(["", "", "", f"{labels['nis']} {client['nis']}"])
    if settings.get("nif"):
        sheet.append([f"{labels['nif']} {settings.get('nif', '')}", "", "", ""])
    if settings.get("registrationNumber"):
        sheet.append([f"{labels['registrationNumber']} {settings.get('registrationNumber', '')}", "", "", ""])
    if settings.get("rip"):
        sheet.append([f"{labels['rip']} {settings.get('rip', '')}", "", "", ""])
    sheet.append([])

    headers = [labels["description"], labels["quantity"], labels["unitPrice"], labels["amount"]]
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.font = bold
        cell.fill = header_fill

    for item in invoice.get("items", []):
        sheet.append(
            [
                item["description"],
                item["quantity"],
                item["unitPrice"],
                item["amount"],
            ]
        )

    sheet.append([])
    totals_start_row = sheet.max_row + 1
    sheet.append(["", "", labels["subtotal"], totals.get("subtotal", 0)])
    sheet.append(["", "", labels["adjustments"], totals.get("adjustment", 0)])
    sheet.append(["", "", labels["discount"], -totals.get("discount", 0)])
    sheet.append(["", "", labels["adjustedSubtotal"], totals.get("taxableSubtotal", 0)])
    sheet.append(["", "", f"{labels['tax']} ({invoice.get('taxRate', 0):g}%)", totals.get("tax", 0)])
    sheet.append(["", "", labels["total"], totals.get("total", 0)])
    sheet.append([])
    sheet.append([labels["notes"], invoice.get("notes") or settings.get("footerNotes") or ""])

    for row in sheet.iter_rows(min_row=totals_start_row, max_row=totals_start_row + 5, min_col=3, max_col=4):
        for cell in row:
            cell.alignment = right
            if cell.column == 4:
                cell.number_format = "#,##0.00"

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 14
    sheet.column_dimensions["E"].width = 14

    if invoice.get("includeCachet") and settings.get("cachetImage"):
        cachet_bytes = decode_logo_image(settings["cachetImage"])
        if cachet_bytes:
            xl_img = XLImage(io.BytesIO(cachet_bytes))
            xl_img.width = 112
            xl_img.height = 112
            cachet_row = totals_start_row + 9
            sheet.add_image(xl_img, f"D{cachet_row}")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()



